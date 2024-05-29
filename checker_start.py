import os
import sys

import time
import json
import pickle

import pymysql
import requests
import telebot
import ua_generator
from concurrent.futures import ThreadPoolExecutor
from openpyxl import load_workbook
from bs4 import BeautifulSoup

from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from config import host, db_name, user, password

path_to_dir = os.path.dirname(sys.executable)

def get_path(name_path):
    con = pymysql.connect(
        host=host,
        port=3306,
        user=user,
        password=password,
        database=db_name,
        cursorclass=pymysql.cursors.DictCursor
    )

    with con.cursor() as cursor:
        check_all_rows = f"SELECT * FROM Checker_paths"
        cursor.execute(check_all_rows)
        data_sql = cursor.fetchall()
        for data in data_sql:
            if data['name_path'] == f'{name_path}':

                path = [
                    (By.XPATH, data['value_xpath']),
                    (By.CSS_SELECTOR, data['value_css']),
                    (By.CLASS_NAME, data['value_class'])
                ]
                return path


google_btn = get_path('google_btn')
accept_auth_btn_path = get_path('accept_auth_btn_path')
auth_with_sberid_path = get_path('auth_with_sberid_path')



google_link_for_mm = 'https://www.google.com/search?q=%D0%BC%D0%B5%D0%B3%D0%B0%D0%BC%D0%B0%D1%80%D0%BA%D0%B5%D1%82&oq=%D0%BC%D0%B5%D0%B3%D0%B0%D0%BC%D0%B0%D1%80%D0%BA%D0%B5%D1%82&gs_lcrp=EgZjaHJvbWUyBggAEEUYOdIBCDI5NDdqMGo0qAIAsAIA&sourceid=chrome&ie=UTF-8'


botEx = telebot.TeleBot('6516240750:AAHHSC0BlT4xloCif5DP-45NoHBVbQ9Ogtk')
proxy_attempts = None
proxy_list = []
proxy_host = None
proxy_port = None
proxy_password = None
proxy_username = None
user_na = ''

def load_cookies(driver, cook_name):
    print('1')
    try:
        with open(f'{path_to_dir}/mainData/smm_auto_check_settings.json', 'r', encoding='utf-8') as f:
            print('2')
            data = json.load(f)
            print('3')
            cookies_way = data['cookies_way']
            print('4')
            use_sber_id_cookies = data['use_sber_id_cookies']
            print('5')

            # >>> Загрузка JSON куков в браузер
            try:
                print('Выгрузка')
                try:
                    with open(f'{cookies_way}/{cook_name}', 'r', encoding='utf-8') as f:
                        cookies = json.load(f)
                except json.JSONDecodeError as ex:
                    print('Это гологин епта')
                    if str(ex) == 'Unexpected UTF-8 BOM (decode using utf-8-sig): line 1 column 1 (char 0)':
                        with open(f'{cookies_way}/{cook_name}', 'r', encoding='utf-8-sig') as f:
                            cookies = json.load(f)

                # >>> Меняем Lax
                for cookie in cookies:
                    if 'sameSite' in cookie:
                        cookie['sameSite'] = 'Lax'

                # >>> Вытаскиваем токены
                if use_sber_id_cookies == 'True':
                    id_user = None
                    for cookie in cookies:
                        if cookie.get('name') == 'id_user':
                            id_user = cookie
                            cookies_data = id_user
                            break
                else:
                    ecom_token = None
                    for cookie in cookies:
                        if cookie.get('name') == 'ecom_token':
                            ecom_token = cookie
                            cookies_data = ecom_token
                            break
                print(cookies_data)

                with open(f'{cook_name}.pkl', 'wb') as f:
                    pickle.dump(cookies_data, f)

                with open(f'{cook_name}.pkl', 'rb') as f:
                    cookies = pickle.load(f)

                driver.add_cookie(cookies)
                os.remove(f'{cook_name}.pkl')

                print('Выгрузка закончилась')
            except Exception as ex:
                print(ex)
    except Exception as ex:
        print(ex)


def start(cook):
    # >>> Чтение настроек
    with open(f'{path_to_dir}/mainData/smm_auto_check_settings.json', 'r', encoding='utf-8') as f:
        data = json.load(f)
        cookies_way = data['cookies_way']
        use_txt_proxy = data['use_txt_proxy']
        use_send_data_telegram = data['use_send_data_telegram']
        telegram_api = data['telegram_api']
        check_promocode = data['check_promocode']
        check_status_order = data['check_status_order']
        check_sber_bonus = data['check_sber_bonus']
        cookies_uses_way = data['cookies_uses_way']
        use_sber_id_cookies = data['use_sber_id_cookies']
        link_change_mobile_proxy = data['link_change_mobile_proxy']
        use_mobile_proxy = data['use_mobile_proxy']
    print('In Main Func start')

    # >>> Переменные данных об аккаунте
    bonus_value = None
    valut_promocode = None
    order_status = None
    check_order = None

    find_name_cook = cook.split('.')
    cook_name = cook
    cook_type = find_name_cook[1]

    if cook_type == 'txt':
        # Чтение данных из файла и преобразование в JSON
        with open(f'{cookies_way}/{cook_name}', 'r') as file:
            data = file.read()
            json_data = json.loads(data)

        os.remove(f'{cookies_way}/{cook_name}')
        load_txt_cook = cook_name.split('.')
        name_txt_cook = load_txt_cook[0]

        # Запись данных в JSON-файл
        with open(f'{cookies_way}/{name_txt_cook}.json', 'w', encoding='utf-8') as f:
            json.dump(json_data, f, indent=4, ensure_ascii=False)
        cook_name = f'{name_txt_cook}.json'

    # >>> Создание объекта класса webdriver
    def get_chromedriver():
        options = webdriver.ChromeOptions()
        ua = ua_generator.generate(device='desktop', browser='chrome')
        prefs = {"profile.default_content_setting_values.notifications": 2}
        geo = {"profile.default_content_setting_values.geolocation": 2}
        options.add_experimental_option("prefs", prefs)
        options.add_experimental_option("prefs", geo)
        options.add_argument(f'user-agent={ua}')
        options.add_argument("--ignore-certificate-errors")
        options.add_argument("--allow-running-insecure-content")
        options.add_argument("--enable-automation")
        options.add_argument("--disable-infobars")
        options.add_argument("--disable-notifications")
        options.add_argument("--disable-save-password-bubble")
        options.add_argument("--disable-translate")
        options.add_argument("--disable-offer-upload-credit-cards")
        options.add_experimental_option("prefs", prefs)
        options.page_load_strategy = 'eager'
        if use_txt_proxy == 'True':
            PROXY = f'{proxy_host}:{proxy_port}'
            options.add_argument('--proxy-server=%s' % PROXY)
            options.add_extension('proxt_auto_auth.crx')
        else:
            pass
        if use_mobile_proxy == 'True':
            PROXY = f'{proxy_host}:{proxy_port}'
            options.add_argument('--proxy-server=%s' % PROXY)
            options.add_extension('proxt_auto_auth.crx')
        else:
            pass
        driver = webdriver.Chrome(options=options)
        return driver

    # >>> Получение экземпляра driver
    try:
        driver = get_chromedriver()
    except Exception as ex:
        botEx.send_message(882124917, f'Ошибка у пользователя {user_na}\nSMM check\n\n{ex}\nСТРОКА SMM Checker 81 ОШИБКА В НАСТРОЙКЕ БРАУЗЕРА')
        return

    # >>> Переход на google.com
    if use_txt_proxy == 'True' or use_mobile_proxy == 'True':
        driver.get('chrome-extension://ggmdpepbjljkkkdaklfihhngmmgmpggp/options.html')
        tabs = driver.window_handles
        driver.switch_to.window(tabs[0])
        driver.set_window_size(1920, 1080)
        time.sleep(2)
        driver.refresh()
        try:
            input_proxy_login = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.CSS_SELECTOR, '#login')))
            input_proxy_login.send_keys(f'{proxy_username}')
        except Exception as ex:
            print(ex)

        try:
            input_proxy_password = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.CSS_SELECTOR, '#password')))
            input_proxy_password.send_keys(f'{proxy_password}')
        except Exception as ex:
            print(ex)

        try:
            accept_proxy_settings = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.CSS_SELECTOR, '#save')))
            accept_proxy_settings.click()
        except Exception as ex:
            print(ex)
    else:
        pass

    if use_mobile_proxy == 'True':
        while True:
            url = f'{link_change_mobile_proxy}'

            response = requests.get(url)
            if response.status_code == 200:
                print('Поменялся')
                break
            else:
                print('Не поменялся')
    else:
        pass

    driver.refresh()
    time.sleep(0.5)
    driver.refresh()

    if use_mobile_proxy == 'False':
        driver.get('https://www.google.ru/')
        driver.set_window_size(1920, 1080)
        # >>> Подтверждение cookies в google.com
        try:
            tring = "//div[text()='Accept all']"
            trs = WebDriverWait(driver, 2).until(EC.element_to_be_clickable((By.XPATH, tring)))
            trs.click()
        except:
            pass

        # >>> Нажатие на кнопку мне повезет
        for by, seletor in google_btn:
            try:
                cap = WebDriverWait(driver, 5).until(EC.element_to_be_clickable((by, seletor)))
                cap.click()
                driver.get(google_link_for_mm)
            except:
                pass
        # >>> Подтверждение cookies в google.com
        try:
            tring = "//div[text()='Accept all']"
            trs = WebDriverWait(driver, 2).until(EC.element_to_be_clickable((By.XPATH, tring)))
            trs.click()
        except:
            pass

        # >>> Переход на мм
        path = "//*[contains(text(), 'Мегамаркет')]"
        try:
            elements = WebDriverWait(driver, 15).until(EC.element_to_be_clickable((By.XPATH, path)))
            elements.click()
        except:
            pass
        time.sleep(15)
    else:
        driver.get('https://megamarket.ru/')
        driver.set_window_size(1920, 1080)
        time.sleep(15)

    html_check_valid = driver.page_source
    soup_check_valid = BeautifulSoup(html_check_valid, 'html.parser')
    status_elements_order = soup_check_valid.find_all(class_='header-profile-actions')
    statuses_check_valid = [status.get_text(strip=True) for status in status_elements_order]
    status_valid = statuses_check_valid
    print(status_valid)
    if len(status_valid) != 0:
        print('капчи нет')
    else:
        print('Капча')
        driver.close()
        return

    # >>> Нажатие на кнопку авторизоваться
    if use_sber_id_cookies == 'True':
        for by, selector in auth_with_sberid_path:
            try:
                auth_with_sberid = WebDriverWait(driver, 20).until(EC.element_to_be_clickable((by, selector)))
                auth_with_sberid.click()
                break
            except:
                pass
    else:
        pass

    # >>> Нажатие на кнопку войти по сберID
    if use_sber_id_cookies == 'True':
        try:
            entry_number_btn = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.CSS_SELECTOR, '[id^="i-sbid-button-"]')))
            entry_number_btn.click()
        except Exception as ex:
            botEx.send_message(882124917, f'Ошибка у пользователя {user_na}\n{ex}\n СТРОКА 163 SMM Checker')
            print(ex)
    else:
        pass
    print('ВЫЗОВ load_cookies')
    load_cookies(driver, cook_name)
    time.sleep(3)

    # >>> Рефиш страницы для cookies со сберID
    if use_sber_id_cookies == 'True':
        time.sleep(2)
        driver.refresh()
    else:
        pass

    # >>> Нажатие на кнопку продолжить
    if use_sber_id_cookies == 'True':
        for by, selector in accept_auth_btn_path:
            try:
                accept_auth_btn = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((by, selector)))
                accept_auth_btn.click()
                break
            except:
                pass
    else:
        pass

    print('status')

    # >>> Проверка статусов заказаов
    try:
        if check_status_order == 'True':
            """
            Статус заказа
            """
            driver.get('https://megamarket.ru/personal/order/#?orderFilter=ORDER_FILTER_NOT_HIDDEN')
            time.sleep(10)
            driver.get_screenshot_as_file(f'{cook}orderstatus.png')
            html_orders = driver.page_source
            soup_order = BeautifulSoup(html_orders, 'html.parser')
            status_elements_order = soup_order.find_all(class_='order-delivery-status')
            statuses = [status.get_text(strip=True) for status in status_elements_order]
            order_status = statuses
        else:
            order_status = 'Don`t checked'
    except Exception as ex:
        botEx.send_message(882124917, f'Ошибка у пользователя {user_na}\nSMM check\n\n{ex}\nСТРОКА SMM Checker 211 ОШИБКА В ЧЕКЕ СТАТУСОВ ЗАКАЗА')

    print('promocode')

    # >>> Проверка промокодов на аккаунте
    try:
        if check_promocode == 'True':
            """
            Проверка промокодов
            """
            driver.get('https://megamarket.ru/personal/promo-codes')
            time.sleep(10)
            driver.get_screenshot_as_file(f'{cook}prmocode.png')
            html_promocodes = driver.page_source
            soup_promocode = BeautifulSoup(html_promocodes, 'html.parser')
            status_elements_promocode = soup_promocode.find_all(class_='c-button__content')
            promo_values = [status.get_text(strip=True) for status in status_elements_promocode]
            valut_promocode = promo_values
            print(promo_values)
        else:
            valut_promocode = 'Don`t checked'
    except Exception as ex:
        botEx.send_message(882124917, f'Ошибка у пользователя {user_na}\nSMM check\n\n{ex}\nСТРОКА SMM Checker 228 ОШИБКА В ЧЕКЕ ПРОМОКОДОВ')

    print('bonus')

    # >>> Проверка бонусов на аккаунте
    try:
        if check_sber_bonus == 'True':
            """
            Проверка кол-ва бонусов
            """
            driver.get('https://megamarket.ru/personal/loyalty')
            time.sleep(10)
            driver.get_screenshot_as_file(f'{cook}bonus.png')
            html_sber_bonus = driver.page_source
            soup_sber_bonus = BeautifulSoup(html_sber_bonus, 'html.parser')
            status_elements_bonus = soup_sber_bonus.find_all(class_='profile-loyalty__balance-value')
            bonus_value = [status.get_text(strip=True) for status in status_elements_bonus]
            print(bonus_value)
        else:
            bonus_value = 'Don`t checked'
    except Exception as ex:
        botEx.send_message(882124917, f'Ошибка у пользователя {user_na}\nSMM check\n\n{ex}\nСТРОКА SMM Checker 245 ОШИБКА В ЧЕКЕ БОНУСОВ')

    # >>> Запись данных в Excel таблицу
    try:
        wb = load_workbook(f"{path_to_dir}/data.xlsx")
        ws = wb.active
        start_row = 1
        while ws.cell(row=start_row, column=1).value is not None:
            start_row += 1
        column = 1
        if order_status != 'Don`t checked':
            if order_status == []:
                pass
            else:
                order_status = order_status[0]
        else:
            check_order = 'Don`t checked'
        if check_status_order == 'True':
            if len(order_status) == 0:
                check_order = 'Нет заказов'
            else:
                check_order = 'Есть заказы'

        data = [f'{cook}', f'{valut_promocode}', f'{bonus_value}', f'{order_status}', f'{check_order}']
        for item in data:
            ws.cell(row=start_row, column=column, value=item)
            column += 1
        wb.save("data.xlsx")
    except Exception as ex:
        botEx.send_message(882124917, f'Ошибка у пользователя {user_na}\nSMM check\n\n{ex}\nСТРОКА SMM Checker 270 ОШИБКА В ЗАПИСИ ТАБЛИЦЫ')

    # >>> Удление использованных cookies
    try:
        os.remove(f'{cookies_way}/{cook_name}')
    except Exception as ex:
        botEx.send_message(882124917, f'Ошибка у пользователя {user_na}\nSMM check\n\n{ex}\nСТРОКА SMM Checker 272 ОШИБКА В УДАЛЕНИИ ЮЗАНЫХ КУКОВ')

    # >>> Запись проверенных cookies
    try:
        cookies = driver.get_cookies()
        for cookie in cookies:
            if 'sameSite' in cookie:
                cookie['sameSite'] = 'lax'
        with open(f'{cookies_uses_way}/{cook_name}', 'w', encoding='utf-8') as f:
            json.dump(cookies, f)
    except Exception as ex:
        botEx.send_message(882124917, f'Ошибка у пользователя {user_na}\nSMM check\n\n{ex}\nСТРОКА SMM Checker 284 ОШИБКА В СОЗДАНИИ ЧЕКНУТЫХ КУКОВ')

    # >>> Отправка данных в Telegram bot
    if use_send_data_telegram == 'True':
        bot = telebot.TeleBot(telegram_api)
        if os.path.exists(f'{cook}orderstatus.png') == True:
            try:
                a = open(f'{cook}orderstatus.png', 'rb')
                bot.send_photo(int(user_na), a, caption= f'{cook}, Статусы заказов')
                a.close()
                os.remove(f'{cook}orderstatus.png')
            except Exception as ex:
                print(ex)
        else:
            pass
        if os.path.exists(f'{cook}prmocode.png') == True:
            try:
                b = open(f'{cook}prmocode.png', 'rb')
                bot.send_photo(int(user_na), b, caption= f'{cook}, Промокоды')
                b.close()
                os.remove(f'{cook}prmocode.png')
            except Exception as ex:
                print(ex)
        else:
            pass
        if os.path.exists(f'{cook}bonus.png') == True:
           try:
                c = open(f'{cook}bonus.png', 'rb')
                bot.send_photo(int(user_na), c, caption= f'{cook}, Бонусы спасибо')
                c.close()
                os.remove(f'{cook}bonus.png')
           except Exception as ex:
               print(ex)
        else:
            pass
        bot.send_message(int(user_na), f'Аккаунт {cook}\n\nПромокоды: {valut_promocode}\nСтатус заказа: {order_status}\nБонусы спасибо: {bonus_value}\nЕсть ли заказы: {check_order}')
    else:
        pass
    time.sleep(1)
    driver.close()
    return


def Pool(user_n):
    global user_na
    global proxy_list
    user_na = user_n
    with open(f'{path_to_dir}/mainData/smm_auto_check_settings.json', 'r', encoding='utf-8') as f:
        print('asd12312312312')
        data = json.load(f)
        pool_value = data['pool_value']
        use_txt_proxy = data['use_txt_proxy']
        txt_uses_proxy_way = data['txt_uses_proxy_way']
        cookies_way = data['cookies_way']
        txt_proxy_way = data['txt_proxy_way']
        use_mobile_proxy = data['use_mobile_proxy']
        mobile_proxy = data['mobile_proxy']
        cookies_list = os.listdir(f'{cookies_way}')
        if use_txt_proxy == 'True':
            with open(f'{txt_proxy_way}', 'r') as f:
                proxy = f.read()
                proxy_list = proxy.split('\n')
                print(proxy_list)
        else:
            pass

    def process_start(cook):
        start(cook)

    with ThreadPoolExecutor(max_workers=int(pool_value)) as executer:
        print(cookies_list)
        for cook in cookies_list:
            try:
                if use_txt_proxy == 'True':
                    global proxy_attempts, proxy_host, proxy_port, proxy_username, proxy_password
                    if use_txt_proxy == 'True':
                        global proxy_host, proxy_port, proxy_username, proxy_password
                        proxy_get = proxy_list[0]
                        b = proxy_get.split(':')
                        # Настройка прокси
                        proxy_host = b[0]
                        proxy_port = b[1]
                        proxy_username = b[2]
                        proxy_password = b[3]
                        del proxy_list[0]
                        a = open(f'{txt_uses_proxy_way}', 'r', encoding='utf-8')
                        if a == '':
                            with open(f'{txt_uses_proxy_way}', 'w', encoding='utf-8') as f:
                                f.write(proxy + '\n')
                        else:
                            with open(f'{txt_uses_proxy_way}', 'a', encoding='utf-8') as f:
                                f.write(proxy + '\n')
                        a.close()
                    elif use_mobile_proxy == 'True':
                        data = mobile_proxy.split(':')
                        proxy_host = data[0]
                        proxy_port = data[1]
                        proxy_username = data[2]
                        proxy_password = data[3]
                else:
                    pass
            except Exception as ex:
                botEx.send_message(882124917, f'Ошибка у пользователя {user_na}\nSMM check\n\n{ex}\nСТРОКА SMM Cheker 350 ОШИБКА В ПРОКСИ')
                print(ex)
                return
            executer.submit(process_start, cook)
            time.sleep(7)
        executer.shutdown(wait=True)

